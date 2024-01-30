import ai
import json
import app
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
import asyncio

def dict_to_excel(dictionary):
    """
    Convert a nested dictionary to a formatted Excel file.

    Parameters:
    - dictionary (dict): The nested dictionary to be converted.

    Returns:
    - BytesIO: BytesIO object containing the Excel file.
    """
    # Initialize lists to store data for each column
    question_list, response_list, code_list, master_theme_list = [], [], [], []

    # Dictionary to keep track of added questions and their master themes for each master theme
    added_questions = {}

    # Iterate through the dictionary
    for key, value in dictionary.items():
        # Extract data from the dictionary
        master_theme = value.get('Master theme', '')
        question = value.get('question', '')
        responses = value.get('Responses', [])

        # Check if the master theme is in the dictionary
        if master_theme not in added_questions:
            added_questions[master_theme] = set()

        # Check if the question and master theme combination has been added for this master theme
        if (question, master_theme) not in added_questions[master_theme]:
            added_questions[master_theme].add((question, master_theme))

            # Append data to lists
            for response in responses:
                question_list.append(question if len(added_questions[master_theme]) == 1 else '')
                master_theme_list.append(master_theme if len(added_questions[master_theme]) == 1 else '')
                response_list.append(response.get('response', ''))
                code_list.append(response.get('theme', ''))
        print(question_list)

    # Create a DataFrame
    df = pd.DataFrame({
        'Question': question_list,
        'Responses': response_list,
        'Code': code_list,
        'Master Theme': master_theme_list
    })

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with BytesIO() as excel_buffer:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Write the DataFrame to the Excel file
            df.to_excel(writer, sheet_name='Thematic Analysis', index=False)

            # Access the XlsxWriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Thematic Analysis']

            # Set column width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 3)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            # Format the cells
            for row in worksheet.iter_rows(min_row=1, max_col=4, max_row=1):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    cell.font = Font(bold=True)
            for col_index in [1, 4]:
               merge_similar_values(worksheet, column_index=col_index)
            # Format the cells for columns 1 and 4

        # Save the Excel buffer
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

def merge_similar_values(sheet, column_index):
    current_value = None
    start_row = 1

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)

        if cell.value != current_value:
            if start_row < row - 1:
                sheet.merge_cells(start_row=start_row, end_row=row - 1, start_column=column_index, end_column=column_index)
            start_row = row
            current_value = cell.value

    # Merge the last set of cells if needed
    if start_row < sheet.max_row:
        sheet.merge_cells(start_row=start_row, end_row=sheet.max_row, start_column=column_index, end_column=column_index)
   
def convert_array_to_dict(array):
    result_dict = {}
    for index, item in enumerate(array, start=1):
        result_dict[index] = item
    return result_dict

def json_string_to_dict(json_string):
    try:
        # Parse the JSON-formatted string to a dictionary
        json_data = json.loads(json_string)
        return json_data
    except json.JSONDecodeError as e:
        print(f'Error decoding JSON string: {e}')
        return None

def to_array(excel):
    test = []
    for I in excel:
        temp = 'Extract the theme for each response to the question and determine the master or main theme based on the themes (make the mastertheme very specific) and put it into a json format that is structured like this {"question":"","Responses":[{"theme":"","response":""}],"Master theme":""} the question is: ' + f'"{I}" the responses are: '
        responses = ''
        for J in excel[I]:
            responses += f'"{excel[I][J]}", '
        temp += responses
        test.append(temp)
    
    return(test)
#test_dict = json_string_to_dict(ai.prompt('Extract the theme for each response to the question and put it into a json format that is structured like this {"question":"","Responses":[{"theme":"","response":""}]} the question is:' + f'"{}" the responses are: "Yes, sometimes because it covers half of your face. Especifically, when I have acne.", "As I stated about it hides my appearance", "It helps me show my confident-self.", "It doest help me since I dont have social anxiety and Im not self conscious.","It doesnt really help me with my social anxiety since it is just partly covers half of my face and it doesnt hide you from everyone. By self consciousness I think mask are a great help because it can cover my insecurities that i wanted to hide outside from people.","In terms of coping with social anxiety and Self-consciousness, facemask does not really help. Since it only covers a part of my face and doesnt really hide me from individuals.","As it covers half of my face, I know people cant judge me. Hence, it boost my self-confidence.","It helps me to hide my insecurities specially when Im in public places.","It can hide my nose and mouth which is my insecurities."')["choices"][0]["message"]["content"]))