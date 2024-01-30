import ai
import json
import app
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO
import asyncio

import pandas as pd
from openpyxl.styles import Alignment, Font
from io import BytesIO
import os

def export(excel_content, filename='output.xlsx'):
    """
    Save Excel content to a local file.

    Parameters:
    - excel_content (bytes): The Excel file content in bytes.
    - filename (str): The name of the file to save.

    Returns:
    - str: The full path to the saved file.
    """
    # Specify the path to your local directory
    local_directory = 'C:/Users/punci/PVALUE'  # Replace with your actual directory path

    # Ensure the directory exists
    os.makedirs(local_directory, exist_ok=True)

    # Create the full path to the file
    file_path = os.path.join(local_directory, filename)

    # Write the Excel content to the file
    with open(file_path, 'wb') as file:
        file.write(excel_content)

    return file_path


dict = {
   "1":{
      "question":"In what situation do you take off your facemask?",
      "Responses":[
         {
            "theme":"Eating and school routines",
            "response":"Before i eat, after i go to school."
         },
         {
            "theme":"Eating, drinking, and singing activities",
            "response":"Everytime i eat, drink, and sing."
         },
         {
            "theme":"Feeling suffocated and arrival home",
            "response":"While eating, running because I feel suffocated, and when I got home"
         },
         {
            "theme":"Hot weather and sweaty face",
            "response":"When its hot, usually my face is easily to sweat."
         },
         {
            "theme":"Stuffy nose and fresh air needs",
            "response":" Whenever i feel stuffy and needs some fresh air."
         },
         {
            "theme":"Eating and stuffy nose needs",
            "response":" Whenever I am eating or when my face get sweaty."
         },
         {
            "theme":"Facemask irritation and safety level",
            "response":" Personally, I don't wear facemask anymore since feel safe in the school and it irritates my skin the longer I wear it."
         },
         {
            "theme":"Activities and suffocated feeling",
            "response":"It varies in different situations but I mostly take off my mask whenever I eat or in break time but also during our P.E time since I get suffocated sometimes."
         },
         {
            "theme":"Various situations and needs",
            "response":"When there's no one around and when I'm at home. It depends on various situation. Like activities that needs to take my mask off in order to breath properly and when taking pictures. For school purposes."
         }
      ],
      "Master theme":"Individual preferences and needs related to facemask usage in different situations"
   },
   "2":{
      "question":"How does wearing facemask make you feel comfortable?",
      "Responses":[
         {
            "theme":"Protection and Safety",
            "response":"It keeps me safe from the viruses or gems we encounter everytime."
         },
         {
            "theme":"Self-consciousness",
            "response":"Wearing facemask can hide nose and mouth."
         },
         {
            "theme":"Self-confidence",
            "response":"I know that it makes me safe, and it lessen my insecurities"
         },
         {
            "theme":"Discomfort",
            "response":"NO."
         },
         {
            "theme":"Self-consciousness",
            "response":"Its mostly due to it hiding my appearance."
         },
         {
            "theme":"Self-confidence",
            "response":"It boost my confidence as it somehow hides some of my insecurities. Aside from that, it prevent me catching viruses."
         },
         {
            "theme":"Rare comfort",
            "response":" Maybe in some instances, it gives me comfort especially in public transports, but I dont usually wear it. "
         },
         {
            "theme":"Self-awareness",
            "response":"It gives off a sense of comfort since it covers half of my face and I dont need to be insecure about my face or I dont need to be too mindful of my appearance. "
         },
         {
            "theme":"Protection and Safety",
            "response":"It makes me feel comfortable because I know that it can somehow protect me from bacterias that is in the air. It also hides half of my face. "
         }
      ],
      "Master theme":"Protection and Safety"
   },
   "3":{
      "question":"Overall, how does the practice of wearing a facemask affect your daily life?",
      "Responses":[
         {
            "theme":"Benefits and drawbacks",
            "response":"Sometimes it is very useful, however wearing it in our everyday lifes sometimes it irritates me because of the weather. But then, it is useful to prevent having a disease or contagious."
         },
         {
            "theme":"Effectiveness",
            "response":"Nothing much really, its not too effective."
         },
         {
            "theme":"Self-esteem",
            "response":"Wearing facemask can boost self-esteem and it helps in showing what you truly can do as a new version of yourself."
         },
         {
            "theme":"Return to normalcy",
            "response":"For me, it doesn't affect anything since i think everything was back to the old normal"
         },
         {
            "theme":"Facemask as a tool",
            "response":"Facemask are just a tool to fight diseases and it is great to help fight it but if were speaking about self esteem and etc. It doesnt affect my life. Overall though I think it is great to wear because if you have insecurities in your face you can easily cover your face up."
         },
         {
            "theme":"Health benefits",
            "response":"It affects my life in a good way since it protects me from bacterias that I might get from inhaling the air."
         },
         {
            "theme":"Safety and self-confidence",
            "response":"It helps me keep safe, and it boosts my self-confidence, and lessen my social anxiety."
         },
         {
            "theme":"Avoiding viruses/germs",
            "response":"It helps me to avoid viruses or germs I encounter everytime I go outside."
         },
         {
            "theme":"Insecurity masking",
            "response":"It helps me to hide my insecurities and facemask is keeping me safe."
         }
      ],
      "Master theme":"Benefits and drawbacks of wearing a facemask in daily life"
   }
}

import pandas as pd
from openpyxl.styles import Alignment, Font
from io import BytesIO

def dict_to_excel(dictionary):
    """
    Convert a nested dictionary to a formatted Excel file.

    Parameters:
    - dictionary (dict): The nested dictionary to be converted.

    Returns:
    - BytesIO: BytesIO object containing the Excel file.
    """
    # Initialize lists to store data for each column
    question_list, response_list, code_list, master_theme_list = [], [], [], []

    # Dictionary to keep track of added questions and their master themes for each master theme
    added_questions = {}

    # Iterate through the dictionary
    for key, value in dictionary.items():
        # Extract data from the dictionary
        master_theme = value.get('Master theme', '')
        question = value.get('question', '')
        responses = value.get('Responses', [])

        # Check if the master theme is in the dictionary
        if master_theme not in added_questions:
            added_questions[master_theme] = set()

        # Check if the question and master theme combination has been added for this master theme
        if (question, master_theme) not in added_questions[master_theme]:
            added_questions[master_theme].add((question, master_theme))

            # Append data to lists
            for response in responses:
                question_list.append(question if len(added_questions[master_theme]) == 1 else '')
                master_theme_list.append(master_theme if len(added_questions[master_theme]) == 1 else '')
                response_list.append(response.get('response', ''))
                code_list.append(response.get('theme', ''))
        print(question_list)

    # Create a DataFrame
    df = pd.DataFrame({
        'Question': question_list,
        'Responses': response_list,
        'Code': code_list,
        'Master Theme': master_theme_list
    })

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with BytesIO() as excel_buffer:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Write the DataFrame to the Excel file
            df.to_excel(writer, sheet_name='Thematic Analysis', index=False)

            # Access the XlsxWriter workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Thematic Analysis']

            # Set column width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 3)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            # Format the cells
            for row in worksheet.iter_rows(min_row=1, max_col=4, max_row=1):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    cell.font = Font(bold=True)
            for col_index in [1, 4]:
               merge_similar_values(worksheet, column_index=col_index)
            # Format the cells for columns 1 and 4

        # Save the Excel buffer
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

def merge_similar_values(sheet, column_index):
    current_value = None
    start_row = 1

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)

        if cell.value != current_value:
            if start_row < row - 1:
                sheet.merge_cells(start_row=start_row, end_row=row - 1, start_column=column_index, end_column=column_index)
            start_row = row
            current_value = cell.value

    # Merge the last set of cells if needed
    if start_row < sheet.max_row:
        sheet.merge_cells(start_row=start_row, end_row=sheet.max_row, start_column=column_index, end_column=column_index)


export(dict_to_excel(dict))